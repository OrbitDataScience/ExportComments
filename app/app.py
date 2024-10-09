import streamlit as st
import requests
from exportcomments import ExportComments
import time
import sys
import openpyxl
import os
import base64
import streamlit.components.v1 as components
import pkg_resources
import random
import string
import datetime
# Defina seu token de API
api_token = 'b11ee661080db564ced715d0f6a88c9adfdbec4e3e7db118f72e720c20defa3b04674c81554a874f8eeba296a0399b2645b34d473fe80eccc5b0a11d'
ex = ExportComments(api_token)

# <Front end>
st.set_page_config(page_title="Export Comments", page_icon="🔗", layout="wide")

st.header('Inserir os Links', divider='rainbow')

with st.form(key="my_form"):
    st.text_area("Insira os links aqui", height=400, max_chars=None, key="url")
    submitted = st.form_submit_button('Extrair Comentários')
# </Front end>

def get_response(guid):
    while True:
        response = ex.exports.check(guid=guid)
        status = response.body['data'][0]['status']    

        if status == 'done':
            break
        elif status == 'error':
            break
            # sys.exit()
            

        time.sleep(5)

    download_url = response.body['data'][0]['downloadUrl']
    headers = {
        'Authorization': api_token,
        'Content-Type': 'application/json',
        'User-Agent': 'python-sdk-{}'.format(pkg_resources.get_distribution('exportcomments').version),
    }

    response = requests.get("https://exportcomments.com/" + download_url, headers=headers)
   
    if response.status_code == 200:
        temp_filename = "result.xlsx"
        with open(temp_filename, "wb") as output:
            output.write(response.content)

        st.success(f"Comentários extraídos com sucesso da URL: {download_url}")
        return temp_filename 
        
    else:
        st.error(f"Erro: {response.status_code} - URL: {download_url}")
        return None

def append_to_excel(main_filename, temp_filename):
    # Carrega o arquivo principal
    main_wb = openpyxl.load_workbook(main_filename)
    main_sheet = main_wb.active

    # Carrega o arquivo temporário
    temp_wb = openpyxl.load_workbook(temp_filename)
    temp_sheet = temp_wb.active

    # Encontra a última linha não vazia no arquivo principal
    main_last_row = main_sheet.max_row

    # Copia os dados do arquivo temporário para o arquivo principal
    for row in temp_sheet.iter_rows(min_row=2):  # Assume que a primeira linha é cabeçalho
        values = [cell.value for cell in row]
        hyperlinks = [cell.hyperlink.target if cell.hyperlink else None for cell in row]
        main_sheet.append(values)
        for col_num, hyperlink in enumerate(hyperlinks, 1):
            if hyperlink:
                main_sheet.cell(row=main_sheet.max_row, column=col_num).hyperlink = hyperlink
                main_sheet.cell(row=main_sheet.max_row, column=col_num).style = "Hyperlink"

    # Salva o arquivo principal
    main_wb.save(main_filename)
    print(f"Dados do arquivo {temp_filename} adicionados ao arquivo {main_filename}")
    os.remove(temp_filename)  # Remove o arquivo temporário após uso

def generate_random_filename():
    current_date = datetime.datetime.now().strftime("%Y%m%d")
    random_string = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    filename = f"excel_file_{current_date}_{random_string}.xlsx"
    return filename


if __name__ == '__main__':
    try:
        if submitted:

            main_filename = generate_random_filename()
    
            # Cria um novo arquivo Excel se não existir
            if not os.path.isfile(main_filename):
                wb = openpyxl.Workbook()
                wb.save(main_filename)

            urls = st.session_state['url'].strip().split('\n')  # Divide os links por linha
            qtd_links = len(urls)
            st.info(f"Quantidade de links: " + str(qtd_links))
            minutos_estimado = qtd_links * 2
            st.info(f"⏳ Iniciando extração de comentários. Tempo estimando: {int(minutos_estimado)} minutos")
            
            count = 0
            for url in urls:
                url = url.strip()  # Remove espaços em branco ao redor da URL
                if not url:
                    continue  # Pula URLs vazias

                try:
                    response = ex.exports.create(
                        url=url, replies='false', twitterType=None
                    )
                    guid = response.body['data']['guid']
                   
                    temp_filename  = get_response(guid)                       

                    if temp_filename is not None:	
                        count += 1
                        append_to_excel(main_filename, temp_filename)
                except Exception as e:
                    st.error(f"Erro ao processar a URL {url}")
            
            st.info(f"{count} URLs foram extraídas com sucesso!")

            # Fornece o arquivo para download automaticamente
            if count > 0:
                with open(main_filename, "rb") as f:
                    bytes_data = f.read()
                    b64 = base64.b64encode(bytes_data).decode()
                    href = f'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}'
                    download_script = f'''
                    <html>
                        <body>
                            <a id="download_link" href="{href}" download="{main_filename}"></a>
                            <script>
                                document.getElementById('download_link').click();
                            </script>
                        </body>
                    </html>
                    '''
                    components.html(download_script)


    except Exception as e:
        st.error(f"Erro: {e}")
